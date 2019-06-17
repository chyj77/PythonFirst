# coding=gbk

import openpyxl
import pymysql
import time

# 生产
DB = pymysql.connect(host='rm-bp1d11o4l8p0u8923bo.mysql.rds.aliyuncs.com', port=3306, user='chenyongjin',
                     password='AsDCDikb97', db='commission',
                     charset='utf8mb4', cursorclass=pymysql.cursors.DictCursor)
# uat
# DB = pymysql.connect(host='bixuan-uat.bisinuolan.cn', port=3306, user='mysql',
#                      password='dev_data2018', db='commission',
#                      charset='utf8mb4', cursorclass=pymysql.cursors.DictCursor)
# 测试
# DB = pymysql.connect(host='172.16.6.71', port=3306, user='mysql',
#                      password='dev_data2018', db='commission',
#                      charset='utf8mb4', cursorclass=pymysql.cursors.DictCursor)

path = r"C:\Users\hailxie\Documents\WeChat Files\chyj_1977\FileStorage\File\2019-05\根据赠品订单号需要匹配出的原始订单号.xlsx"


def readexcel():
    cursor = DB.cursor()
    inwb = openpyxl.load_workbook(path)  # 读文件
    sheetnames = inwb.get_sheet_names()  # 获取读文件中所有的sheet，通过名字的方式
    sheet = inwb.get_sheet_by_name('Sheet1')  # 获取第一个sheet内容

    # 获取sheet的最大行数和列数
    rows = sheet.max_row + 1
    cols = sheet.max_column
    print("最大行数=", sheet.max_row)
    print("最大列数=", cols)
    for r in range(2, rows):
        persentorderno = sheet.cell(r, 2).value
        print(persentorderno)
        sql = "select order_id from bsnl.order_info where order_no='%s'" % (persentorderno)
        print(sql)
        cursor.execute(sql)
        persent = cursor.fetchone();
        if persent is None :
            sheet.cell(row=r, column=3).value = ''
        else:
            order_id = persent['order_id']
            ordersql = "select order_no from bsnl.order_info a,bsnl.partner_order_log b " \
                       "where a.order_id =b.order_id and  b.give_content='%s'" % (order_id)
            print(ordersql)
            cursor.execute(ordersql)
            orderno = cursor.fetchone();
            if orderno is None :
                sheet.cell(row=r, column=3).value = ''
            else :
                sheet.cell(row=r, column=3).value = orderno['order_no']

    inwb.save(filename=path)

if __name__ == '__main__':
    begin = time.time()
    beginTime = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(time.time()))
    readexcel()
    total = (time.time() - begin)
    endTime = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(time.time()))
    m, s = divmod(total, 60)
    h, m = divmod(m, 60)
    print('开始时间：', beginTime)
    print('结束时间：', endTime, " 耗时:", "%02d 小时 %02d 分钟 %02d 秒" % (h, m, s))

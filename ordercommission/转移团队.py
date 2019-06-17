# -*- coding: utf-8 -*-

from openpyxl import Workbook
import pymysql
import time

# 生产
DB = pymysql.connect(host='rm-bp1d11o4l8p0u8923bo.mysql.rds.aliyuncs.com', port=3306, user='chenyongjin',
                     password='AsDCDikb97', db='commission',
                     charset='utf8mb4', cursorclass=pymysql.cursors.DictCursor)
# uat
# DB = pymysql.connect(host='bixuan-uat.bisinuolan.cn', port=3306, user='mysql',
#                      password='dev_data2018', db='commission',
#                      charset='utf8mb4', cursorclass=pymysql.cursors.DictCursor)
# 测试
# DB = pymysql.connect(host='172.16.6.71', port=3306, user='mysql',
#                      password='dev_data2018', db='commission',
#                      charset='utf8mb4', cursorclass=pymysql.cursors.DictCursor)

ximing = '13586777175'
mobile = '13586777175'
oldmobile = '15917421638'
newmobile = '18022903507'

path = r"D:\碧斯诺兰\团队变更\推荐人更改申请%s.xlsx"%ximing
# path = r"C:\Users\hailxie\Desktop\推荐人更改申请%s.xlsx" % ximing


def readexcel():
    cursor = DB.cursor()
    # inwb = openpyxl.load_workbook(path)  # 读文件

    inwb = Workbook()
    # sheetnames = inwb.get_sheet_names()  # 获取读文件中所有的sheet，通过名字的方式
    # sheet = inwb.get_sheet_by_name(sheetnames[0])  # 获取第一个sheet内容
    sheet = inwb.active

    sql = "select * from bsnl.user_info where mobile='%s'" % mobile
    cursor.execute(sql)
    userdata = cursor.fetchone()
    deleFlag = userdata['delete_flag']
    id = userdata['id']
    directorid = userdata['director_id']
    redirectorid = userdata['recommend_director_user_id']
    reareaid = userdata['recommend_area_user_id']
    if deleFlag != 1:
        print('该用户已被禁用')
        return

    oldsql = "select * from bsnl.user_info where mobile='%s'" % oldmobile
    cursor.execute(oldsql)
    olduserdata = cursor.fetchone()
    oldid = olduserdata['id']
    olduserid = olduserdata['user_id']
    newsql = "select * from bsnl.user_info where mobile='%s'" % newmobile
    cursor.execute(newsql)
    newuserdata = cursor.fetchone()
    newid = newuserdata['id']
    newuserid = newuserdata['user_id']
    newredirectorid = newuserdata['recommend_director_user_id']
    newredirectoruserid = newuserdata['recommend_director_userId']
    role = userdata['role']
    if role == 3:
        if redirectorid == oldid and redirectorid != 0:
            role3sql = "select * from bsnl.user_info where id='%s' or director_id='%s' or recommend_director_user_id='%s'" % (
                id, id, id)
            cursor.execute(role3sql)
            role3userdatas = cursor.fetchall()
            r = 2
            for role3userdata in role3userdatas:
                c = 1
                for key, values in role3userdata.items():
                    if r == 2:
                        sheet.cell(row=1, column=c).value = key
                        sheet.cell(row=r, column=c).value = values
                    else:
                        sheet.cell(row=r, column=c).value = values
                    c = c + 1
                r = r + 1
            role3updatesql = "update bsnl.user_info set recommend_director_user_id=%s,\
                        recommend_director_userId='%s' where id='%s' or director_id='%s' or recommend_director_user_id='%s'" % (
                newid, newuserid, id, id, id)
            print(role3updatesql)
        else:
            print(oldmobile + ' 不是 ' + mobile + " 的直招董事！")
            return

    if role == 2:
        if reareaid > 0:
            print(oldmobile + '  有推荐大区，不能转移')
            return
        if directorid == oldid and directorid != 0:
            role2sql = "select * from bsnl.user_info where id='%s' or area_id='%s' or recommend_area_user_id='%s'" % (
                id, id, id)
            cursor.execute(role2sql)
            role2userdatas = cursor.fetchall()
            r = 2
            for role2userdata in role2userdatas:
                c = 1
                for key, values in role2userdata.items():
                    if r == 2:
                        sheet.cell(row=1, column=c).value = key
                        sheet.cell(row=r, column=c).value = values
                    else:
                        sheet.cell(row=r, column=c).value = values
                    c = c + 1
                r = r + 1
            role2updatesql = "update bsnl.user_info set director_id=%s,director_userId='%s' ，recommend_director_user_id=%s,\
                        recommend_director_userId='%s' where id='%s' or area_id='%s' or recommend_area_user_id='%s'" % (
                newid, newuserid, newredirectorid, newredirectoruserid, id, id, id)
            print(role2updatesql)
            children(cursor, sheet, id, newid, newuserid, newredirectorid, newredirectoruserid, r)
        else:
            print(oldmobile + ' 不是 ' + mobile + " 的直招董事！")
            return

    inwb.save(filename=path)


def children(cursor, sheet, id, newid, newuserid, newredirectorid, newredirectoruserid, r):
    r = r + 1
    role2sql = "select * from bsnl.user_info where role=2 and recommend_area_user_id='%s'" % id
    cursor.execute(role2sql)
    role2userdatas = cursor.fetchall()
    if role2userdatas is not None:
        for role2userdata in role2userdatas:
            childid = role2userdata['id']
            childrole2sql = "select * from bsnl.user_info where id='%s' or area_id='%s' or recommend_area_user_id='%s'" % (
                childid, childid, childid)
            cursor.execute(childrole2sql)
            childrole2userdatas = cursor.fetchall()
            for childrole2userdata in childrole2userdatas:
                c = 1
                for key, values in childrole2userdata.items():
                    if r == 2:
                        sheet.cell(row=r, column=c).value = values
                    else:
                        sheet.cell(row=r, column=c).value = values
                    c = c + 1
                r = r + 1
            childrole3updatesql = "update bsnl.user_info set director_id=%s,director_userId='%s' ，recommend_director_user_id=%s,\
                        recommend_director_userId='%s' where id='%s' or area_id='%s' or recommend_area_user_id='%s'" % (
                newid, newuserid, newredirectorid, newredirectoruserid, childid, childid, childid)
            print(childrole3updatesql)
            children(cursor, sheet, childid, newid, newuserid, newredirectorid, newredirectoruserid, r)


if __name__ == '__main__':
    begin = time.time()
    beginTime = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(time.time()))
    print('开始时间：', beginTime)
    readexcel()
    total = (time.time() - begin)
    endTime = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(time.time()))
    m, s = divmod(total, 60)
    h, m = divmod(m, 60)
    print('结束时间：', endTime, " 耗时:", "%02d 小时 %02d 分钟 %02d 秒" % (h, m, s))

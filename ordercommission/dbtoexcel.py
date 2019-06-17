# coding=gbk

import pymysql
import time
import os
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook

# 生产
DB = pymysql.connect(host='rm-bp1d11o4l8p0u8923bo.mysql.rds.aliyuncs.com', port=3306, user='chenyongjin',
                     password='AsDCDikb97', db='commission',
                     charset='utf8mb4', cursorclass=pymysql.cursors.DictCursor)
# uat
# DB = pymysql.connect(host='bixuan-uat.bisinuolan.cn', port=3306, user='mysql',
#                      password='dev_data2018', db='commission',
#                      charset='utf8mb4', cursorclass=pymysql.cursors.DictCursor)
# 测试
# DB = pymysql.connect(host='172.16.6.71', port=3306, user='mysql',
#                      password='dev_data2018', db='commission',
#                      charset='utf8mb4', cursorclass=pymysql.cursors.DictCursor)

path = r"C:\Users\hailxie\Downloads\卡芙琳董事赛事奖励.xlsx"

def getData():
    cursor = DB.cursor();
    sql="SELECT  y.real_name `name`,y.mobile mobile,sum(price) personal,sum(teamprice) team \
            from( SELECT (case when c.role in (0,1,2) then c.director_id else c.id end) id,\
                sum(case when c.role = 3 then og.pay_price * og.goods_number * IFNULL( og.package_number, 1 ) else 0 end) price,\
                 sum( og.pay_price * og.goods_number * IFNULL( og.package_number, 1 ) ) teamprice, c.role \
            FROM bsnl.order_info oi \
            JOIN bsnl.order_goods og ON oi.order_id = og.order_id \
            JOIN bsnl.user_info c ON oi.user_id = c.user_id \
            WHERE 1 = 1  AND og.present_flag <> 1 \
            AND og.goods_id IN (  \
              select goods_id from bsnl.goods where serial_no in ('BBBD950039',\
'BBBD920081',\
'BBBD920082',\
'BBBD910013',\
'BBBD950047',\
'BBBD920103',\
'BBBD920101',\
'BBBD920105',\
'BBBD950046',\
'BBBD920106',\
'BBBD950045',\
'BBBD950043',\
'BBBD920104',\
'BBBD920102',\
'BBBD920100',\
'BBBD950048',\
'BBBD950041',\
'BBBD950044',\
'BBBD950042',\
'BBBD950040')\
            )\
            AND oi.`status` NOT IN ( 1, 5, 6, 7 )\
            AND oi.created_at >= '2019-06-05 10:00:00'\
            AND oi.created_at < '2019-06-11 00:00:00'\
            GROUP BY (case when c.role in (0,1,2) then c.director_id else c.id end))x\
            join bsnl.user_info y on x.id = y.id \
            group by x.id\
            order by team desc"
    cursor.execute(sql)
    print(sql)
    datas = cursor.fetchall()

    wb = load_workbook(path)
    sheet = wb.get_sheet_by_name('Sheet2')
    maxRows = sheet.max_row
    for srow in range(maxRows,-1,-1) :
        sheet.delete_rows(srow)

    tableTitle = ['姓名', '手机号码', '个人业绩', '团队业绩']

    # 维护表头
    #        if row < 1 or column < 1:
    #          raise ValueError("Row or column values must be at least 1")
    # 如上，openpyxl 的首行、首列 是 （1,1）而不是（0,0），如果坐标输入含有小于1的值，提示 ：Row or column values must be at least 1，即最小值为1.
    for col in range(len(tableTitle)):
        c = col + 1
        sheet.cell(row=1, column=c).value = tableTitle[col]

    # 数据表基本信息
    # tableValues = [['张学友', 15201062100, 18, '测试数据！'], ['李雷', 15201062598, 19, '测试数据！'],['Marry', 15201062191, 28, '测试数据！']]

    row = 2
    for data in datas :
        sheet.cell(row, column=1).value = data['name']
        sheet.cell(row, column=2).value = data['mobile']
        sheet.cell(row, column=3).value = data['personal']
        sheet.cell(row, column=4).value = data['team']
        row = row + 1

    #wb.save(ExcelFullName)
    wb.save(filename=path)


if __name__ == '__main__':
    getData()